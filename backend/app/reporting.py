from __future__ import annotations

import csv
import json
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Fixed severity bands (methodology): r = anomaly_score / operating_threshold
SEVERITY_RATIO_HIGH_MIN = 3.0
SEVERITY_RATIO_MEDIUM_MIN = 1.5

def _load_json_if_exists(path: Path) -> Dict[str, Any]:
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}

def _channel_summary_rows(channel_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for cr in channel_results:
        out.append(
            {
                "channel": cr.get("channel_name"),
                "windows": cr.get("num_windows"),
                "anomalies": cr.get("num_anomalies"),
                "anomaly_rate": cr.get("anomaly_rate"),
                "threshold": cr.get("threshold"),
                "results_path": cr.get("results_path"),
            }
        )
    return out

@dataclass
class Finding:
    finding_id: str
    channel: str
    window_index: Optional[int]
    score: float
    threshold: float
    ratio: float  # score / threshold (same basis as severity bands)
    severity: str
    band_rule: str  # which numeric rule produced severity (e.g. "r ≥ 3.0")
    description: str
    possible_cause: str
    recommended_action: str

def _safe_float(x, default=0.0) -> float:
    try:
        return float(x)
    except Exception:
        return float(default)

def _score_threshold_ratio(score: float, threshold: float) -> float:
    thr = max(float(threshold), 1e-9)
    return float(score) / thr


def _severity_from_score(score: float, threshold: float) -> str:
    r = _score_threshold_ratio(score, threshold)
    if r >= SEVERITY_RATIO_HIGH_MIN:
        return "High"
    if r >= SEVERITY_RATIO_MEDIUM_MIN:
        return "Medium"
    if score >= float(threshold):
        return "Low"
    return "Normal"


def _normalize_severity_label(raw: Optional[str], score: float, threshold: float) -> str:
    """
    Prefer severity from analyzer CSV (same labels as /runs/{id}/results UI).
    Falls back to ratio bands if column missing.
    """
    if raw is not None and str(raw).strip() != "":
        s = str(raw).strip().upper()
        mapping = {"NORMAL": "Normal", "LOW": "Low", "MEDIUM": "Medium", "HIGH": "High"}
        if s in mapping:
            return mapping[s]
        s2 = str(raw).strip().title()
        if s2 in ("Low", "Medium", "High", "Normal"):
            return s2
    return _severity_from_score(score, threshold)


def _band_rule_label(severity: str) -> str:
    """Human-readable rule text for the severity band (fixed methodology)."""
    if severity == "High":
        return f"r ≥ {SEVERITY_RATIO_HIGH_MIN:g}"
    if severity == "Medium":
        return f"{SEVERITY_RATIO_MEDIUM_MIN:g} ≤ r < {SEVERITY_RATIO_HIGH_MIN:g}"
    if severity == "Low":
        return f"1 < r < {SEVERITY_RATIO_MEDIUM_MIN:g}"
    return f"r ≤ 1 (not above threshold)"


def _finding_fields_from_measurements(
    channel: str,
    window_index: Optional[int],
    score: float,
    threshold: float,
    severity: str,
    ratio: float,
) -> Tuple[str, str, str]:
    """
    Per-window text derived from this run's measurements and the fixed ratio bands.
    Avoids repeating generic narrative; possible_cause defers to methodology.
    """
    wi = str(window_index) if window_index is not None else "—"
    rule = _band_rule_label(severity)
    desc = (
        f"score={score:.6g}, threshold={threshold:.6g}, r={ratio:.4f} → {severity} "
        f"({rule}). Channel {channel}, window {wi}."
    )
    cause = ""
    rec = (
        f"{severity} ({rule}) — follow mission policy for this band; "
        f"channel={channel}, window={wi}, r={ratio:.4f}."
    )
    return desc, cause, rec

def _overall_state(anomaly_rate: float, high_count: int, medium_count: int) -> str:
    if high_count > 0 or anomaly_rate >= 0.10:
        return "Critical"
    if medium_count > 0 or anomaly_rate >= 0.03:
        return "Suspicious"
    return "Normal"


def _mission_posture_rationale_line(posture: str, summary: Dict[str, Any]) -> str:
    """
    One-line executive explanation of mission posture (no internal rule/debug phrasing).
    Wording follows the same policy as _overall_state, using measured counts and rates.
    """
    ar_pct = float(summary.get("anomaly_rate") or 0) * 100.0
    hi = int(summary.get("high_count") or 0)
    med = int(summary.get("medium_count") or 0)
    anom = int(summary.get("anomaly_count") or 0)
    p = str(posture).strip()

    if p == "Critical":
        has_hi = hi > 0
        high_rate = ar_pct >= 10.0
        if has_hi and high_rate:
            return (
                "Mission posture is classified as Critical due to sustained high-severity anomaly "
                "activity and an elevated aggregate anomaly rate relative to mission tolerance."
            )
        if has_hi:
            return (
                "Mission posture is classified as Critical due to high-severity anomaly windows "
                "relative to the operating threshold."
            )
        if high_rate:
            return (
                "Mission posture is classified as Critical due to an aggregate anomaly rate that "
                "exceeds mission tolerance for routine operations."
            )
        return (
            "Mission posture is classified as Critical based on aggregate anomaly burden for this "
            "telemetry scope."
        )

    if p == "Suspicious":
        has_med = med > 0
        susp_rate = ar_pct >= 3.0
        if has_med and susp_rate:
            return (
                "Mission posture is assessed as Suspicious due to medium-severity anomaly patterns "
                "and an aggregate rate above typical operating norms."
            )
        if has_med:
            return (
                "Mission posture is assessed as Suspicious due to medium-severity anomaly indicators "
                "in this telemetry sample."
            )
        if susp_rate:
            return (
                "Mission posture is assessed as Suspicious due to an aggregate anomaly rate above "
                "routine operating expectations."
            )
        if anom > 0:
            return (
                "Mission posture is assessed as Suspicious based on the severity mix observed across "
                "flagged windows."
            )
        return "Mission posture is assessed as Suspicious based on aggregate indicators for this run."

    return (
        "Mission posture is assessed as Normal: anomaly burden and severity distribution remain "
        "within expected operating bounds for this configuration."
    )


def _interpret_patterns(findings: List[Finding]) -> Dict[str, Any]:
    """
    Lightweight expert-style interpretation based on severity mix and score ratios.
    Does NOT inspect raw model internals. Keeps the system unsupervised.
    """
    def _sx(s: str) -> str:
        return str(s).strip().lower()

    hi = sum(1 for f in findings if _sx(f.severity) == "high")
    med = sum(1 for f in findings if _sx(f.severity) == "medium")
    low = sum(1 for f in findings if _sx(f.severity) == "low")

    # Heuristics (documented as "signals", not ground truth)
    drift = (med + hi) >= 5
    freeze = hi >= 2  # strong anomalies repeatedly
    noise_spike = (low + med) >= 10 and hi == 0
    pattern_shift = hi >= 1 and med >= 2

    return {
        "drift_suspected": bool(drift),
        "freeze_behavior_suspected": bool(freeze),
        "noise_spike_suspected": bool(noise_spike),
        "pattern_shift_suspected": bool(pattern_shift),
        "notes": [
            "Signals use counts of windows in each ratio band (High/Medium/Low) defined in this report.",
            "Confirm with mission events and telemetry context before escalation.",
        ],
    }

def parse_run_findings(channel_results: List[Dict[str, Any]], top_k: int = 25) -> Tuple[Dict[str, Any], List[Finding]]:
    """
    Parse per-channel results CSVs written by analyze pipeline.
    Returns (summary, top_findings).
    """
    total_windows = 0
    anomaly_count = 0
    low = med = high = 0
    threshold_used = None

    # score distribution across ALL windows (not only anomalies)
    score_min = None
    score_max = None
    score_sum = 0.0

    # Split stats: normal windows vs flagged windows (explicit طبيعي / غير طبيعي)
    nw_normal = 0
    nrm_min = nrm_max = None
    nrm_sum = 0.0
    anom_min = anom_max = None
    anom_sum = 0.0

    all_anomalies: List[Finding] = []

    for cr in channel_results:
        ch = str(cr.get("channel_name", "unknown"))
        path = Path(cr.get("results_path", ""))
        if not path.exists():
            continue

        with open(path, "r", encoding="utf-8", newline="") as f:
            r = csv.DictReader(f)
            for row in r:
                score = _safe_float(row.get("score") or row.get("anomaly_score") or 0.0)
                thr = _safe_float(row.get("threshold") or cr.get("threshold") or 0.0)
                is_anom = row.get("is_anomaly")
                try:
                    is_anom_i = int(is_anom) if is_anom is not None else int(score > thr)
                except Exception:
                    is_anom_i = int(score > thr)

                # Same severity labels as analyze pipeline CSV (UI preview uses these rows)
                sev = _normalize_severity_label(row.get("severity"), score, thr)

                total_windows += 1
                if threshold_used is None and thr:
                    threshold_used = float(thr)

                # Global score stats
                if score_min is None or float(score) < float(score_min):
                    score_min = float(score)
                if score_max is None or float(score) > float(score_max):
                    score_max = float(score)
                score_sum += float(score)

                if is_anom_i:
                    if anom_min is None or float(score) < float(anom_min):
                        anom_min = float(score)
                    if anom_max is None or float(score) > float(anom_max):
                        anom_max = float(score)
                    anom_sum += float(score)

                    if sev == "Normal":
                        sev = "Low"
                    anomaly_count += 1
                    if sev == "High":
                        high += 1
                    elif sev == "Medium":
                        med += 1
                    else:
                        low += 1

                    wi = row.get("window_index")
                    window_index = None
                    try:
                        window_index = int(wi) if wi is not None else None
                    except Exception:
                        window_index = None

                    ratio = _score_threshold_ratio(score, thr)
                    rule = _band_rule_label(sev)
                    desc, cause, rec = _finding_fields_from_measurements(
                        ch, window_index, float(score), float(thr), sev, ratio
                    )

                    all_anomalies.append(
                        Finding(
                            finding_id=f"F-{uuid.uuid4().hex[:8]}",
                            channel=ch,
                            window_index=window_index,
                            score=float(score),
                            threshold=float(thr),
                            ratio=float(ratio),
                            severity=str(sev),
                            band_rule=rule,
                            description=desc,
                            possible_cause=cause,
                            recommended_action=rec,
                        )
                    )
                else:
                    nw_normal += 1
                    if nrm_min is None or float(score) < float(nrm_min):
                        nrm_min = float(score)
                    if nrm_max is None or float(score) > float(nrm_max):
                        nrm_max = float(score)
                    nrm_sum += float(score)

    all_anomalies.sort(key=lambda f: f.ratio, reverse=True)
    top_findings = all_anomalies[: max(1, int(top_k))]

    normal_count = max(0, total_windows - anomaly_count)
    anomaly_rate = float(anomaly_count / max(1, total_windows))
    normal_rate = float(normal_count / max(1, total_windows))
    severity_summary = _overall_state(anomaly_rate, high, med)
    score_mean = (score_sum / max(1, total_windows)) if total_windows else None

    nrm_mean = (nrm_sum / max(1, nw_normal)) if nw_normal else None
    anom_mean = (anom_sum / max(1, anomaly_count)) if anomaly_count else None

    summary = {
        "total_windows": int(total_windows),
        "normal_count": int(normal_count),
        "anomaly_count": int(anomaly_count),
        "normal_rate": normal_rate,
        "anomaly_rate": anomaly_rate,
        "low_count": int(low),
        "medium_count": int(med),
        "high_count": int(high),
        "threshold_used": float(threshold_used) if threshold_used is not None else None,
        "score_min": float(score_min) if score_min is not None else None,
        "score_mean": float(score_mean) if score_mean is not None else None,
        "score_max": float(score_max) if score_max is not None else None,
        # Normal windows only (score ≤ threshold / not flagged)
        "normal_windows_score_min": float(nrm_min) if nrm_min is not None else None,
        "normal_windows_score_mean": float(nrm_mean) if nrm_mean is not None else None,
        "normal_windows_score_max": float(nrm_max) if nrm_max is not None else None,
        # Flagged / anomalous windows only
        "anomaly_windows_score_min": float(anom_min) if anom_min is not None else None,
        "anomaly_windows_score_mean": float(anom_mean) if anom_mean is not None else None,
        "anomaly_windows_score_max": float(anom_max) if anom_max is not None else None,
        "severity_summary": severity_summary,
    }
    return summary, top_findings


def read_anomaly_table_rows(channel_results: List[Dict[str, Any]], limit: int = 5000) -> List[Dict[str, Any]]:
    """
    Build a run-level anomaly table from per-channel results CSVs.
    Returns up to `limit` rows total (ordered by score/threshold ratio desc).
    Each row includes: channel, window_index, start, end, score, threshold, decision, severity, explanation.
    """
    rows: List[Dict[str, Any]] = []
    for cr in channel_results:
        ch = str(cr.get("channel_name", "unknown"))
        path = Path(cr.get("results_path", ""))
        if not path.exists():
            continue

        with open(path, "r", encoding="utf-8", newline="") as f:
            r = csv.DictReader(f)
            for row in r:
                score = _safe_float(row.get("score") or row.get("anomaly_score") or 0.0)
                thr = _safe_float(row.get("threshold") or cr.get("threshold") or 0.0)
                sev = _normalize_severity_label(row.get("severity"), score, thr)
                ratio = _score_threshold_ratio(score, thr)

                is_anom = row.get("is_anomaly")
                try:
                    is_anom_i = int(is_anom) if is_anom is not None else int(score > thr)
                except Exception:
                    is_anom_i = int(score > thr)

                decision = "Anomalous" if is_anom_i else "Normal"
                wi = row.get("window_index")
                try:
                    wi_i = int(wi) if wi is not None else None
                except Exception:
                    wi_i = None

                start = row.get("start")
                end = row.get("end")
                try:
                    start_i = int(start) if start not in (None, "", "None") else None
                except Exception:
                    start_i = None
                try:
                    end_i = int(end) if end not in (None, "", "None") else None
                except Exception:
                    end_i = None

                desc, _cause, _rec = _finding_fields_from_measurements(ch, wi_i, score, thr, sev, ratio)
                rows.append(
                    {
                        "channel": ch,
                        "window_index": wi_i,
                        "start_timestep": start_i,
                        "end_timestep": end_i,
                        "score": float(score),
                        "threshold": float(thr),
                        "ratio": float(ratio),
                        "decision": decision,
                        "severity": sev,
                        "explanation": desc,
                    }
                )

    rows.sort(key=lambda x: float(x.get("ratio") or 0.0), reverse=True)
    if limit and len(rows) > int(limit):
        return rows[: int(limit)]
    return rows

def build_report_json(
    *,
    report_id: str,
    run: Dict[str, Any],
    user: Dict[str, Any],
    model_version: str,
    channel_results: List[Dict[str, Any]],
    generated_at: Optional[str] = None,
    model_name: Optional[str] = None,
    model_path_used_for_run: Optional[str] = None,
    threshold_path_used_for_run: Optional[str] = None,
    analysis_meta: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    generated_at = generated_at or datetime.utcnow().isoformat()
    _app_dir = Path(__file__).resolve().parent

    cap = 80
    summary, findings = parse_run_findings(channel_results, top_k=cap)
    findings_in_table = len(findings)
    total_flagged = int(summary.get("anomaly_count") or 0)
    findings_truncated = total_flagged > cap

    interpretation = _interpret_patterns(findings)
    channel_summary = _channel_summary_rows(channel_results)

    # Recommendations: posture from aggregate rules + measured counts/rates for this run
    sev = summary["severity_summary"]
    hi_n = int(summary.get("high_count") or 0)
    med_n = int(summary.get("medium_count") or 0)
    low_n = int(summary.get("low_count") or 0)
    ar_pct = float(summary.get("anomaly_rate") or 0) * 100.0
    if sev == "Critical":
        recs = [
            f"Escalate: High/Med/Low = {hi_n}/{med_n}/{low_n}; rate {ar_pct:.2f}%.",
            "Isolate source if safe; preserve artifacts.",
        ]
    elif sev == "Suspicious":
        recs = [
            f"Review: H/M/L = {hi_n}/{med_n}/{low_n}; rate {ar_pct:.2f}% — correlate with timeline.",
            "Increase monitoring if anomalies persist.",
        ]
    else:
        recs = [
            f"Monitor: H/M/L = {hi_n}/{med_n}/{low_n}; rate {ar_pct:.2f}%.",
        ]

    meta = analysis_meta or {}
    eff_model_name = (model_name or meta.get("model_name") or "").strip() or "—"
    eff_model_path = (model_path_used_for_run or meta.get("model_path") or "").strip() or None
    eff_thr_path = (threshold_path_used_for_run or meta.get("threshold_path") or "").strip() or None
    eff_model_version = (model_version or "").strip() or str(meta.get("model_version") or "").strip() or "—"

    # Use run-specific thresholds/config when available (avoid mismatched static snapshots)
    thresholds_snapshot: Dict[str, Any] = {}
    if eff_thr_path:
        try:
            thresholds_snapshot = _load_json_if_exists(Path(eff_thr_path))
        except Exception:
            thresholds_snapshot = {}
    if not thresholds_snapshot:
        thresholds_snapshot = _load_json_if_exists(_app_dir / "thresholds_qc_filtered.json")
    if not thresholds_snapshot:
        thresholds_snapshot = _load_json_if_exists(_app_dir / "thresholds.json")

    operating_snapshot = _load_json_if_exists(_app_dir / "operating_config.json")
    ot = operating_snapshot.get("operating_threshold") or {}
    thr_used = summary.get("threshold_used")

    # Technical appendix — weights of this report run (same file/analyze as interactive UI when meta exists)
    tech = {
        "model_name": eff_model_name,
        "model_version_used_for_run": eff_model_version,
        "model_weights_path": eff_model_path,
        "threshold_file_path": eff_thr_path,
        "threshold_strategy": "Runtime operating point; numeric threshold from CSV rows.",
        "scoring_formula": "Hybrid score (recon + prediction + gradient); weights from thresholds file.",
        "uploaded_filename": run.get("filename"),
        "file_hash_sha256": run.get("file_sha256"),
        "run_id": run.get("run_id"),
        "generated_report_id": report_id,
        "weights": thresholds_snapshot.get("weights") if isinstance(thresholds_snapshot, dict) else None,
        "available_threshold_keys": list((thresholds_snapshot.get("thresholds") or {}).keys())
        if isinstance(thresholds_snapshot, dict)
        else [],
    }

    file_risk = summary["severity_summary"]
    file_risk_explain = _mission_posture_rationale_line(file_risk, summary)

    report = {
        "cover": {
            "project_name": "CyberSatDetect",
            "report_title": "Telemetry anomaly report",
            "subtitle": "",
            "generated_at": generated_at,
            "analyst": user.get("email") or user.get("sub"),
            "run_id": run.get("run_id"),
            "telemetry_filename": run.get("filename"),
            "document_version": "1.0",
            "classification": "Internal use",
        },
        "executive_summary": {
            "overall_state": summary["severity_summary"],
            "total_windows": summary["total_windows"],
            "normal_windows": summary["normal_count"],
            "anomaly_windows": summary["anomaly_count"],
            "normal_rate": summary["normal_rate"],
            "anomaly_rate": summary["anomaly_rate"],
            "overall_risk_level": summary["severity_summary"],
            "mission_file_assessment": file_risk_explain,
            # طبيعي vs غير طبيعي — نفس الأرقام الموجودة في detection_results مع تسمية واضحة
            "comparison": {
                "normal_windows": {
                    "label_ar": "نوافذ طبيعية (ضمن العتبة)",
                    "label_en": "Normal windows (at or below threshold)",
                    "count": summary["normal_count"],
                    "share": summary["normal_rate"],
                    "score_min": summary["normal_windows_score_min"],
                    "score_mean": summary["normal_windows_score_mean"],
                    "score_max": summary["normal_windows_score_max"],
                },
                "anomaly_windows": {
                    "label_ar": "نوافذ غير طبيعية (مُعلَّمة كشذوذ)",
                    "label_en": "Anomalous windows (flagged above threshold)",
                    "count": summary["anomaly_count"],
                    "share": summary["anomaly_rate"],
                    "score_min": summary["anomaly_windows_score_min"],
                    "score_mean": summary["anomaly_windows_score_mean"],
                    "score_max": summary["anomaly_windows_score_max"],
                },
            },
        },
        "project_scope": {
            "purpose": "AI-assisted anomaly scoring on uploaded telemetry vs an operating threshold.",
            "scope": "Single run — scores from stored channel CSVs; not a substitute for operator judgment.",
        },
        "data_lineage": {
            "inputs": "Upload → cleaning/windowing → model scores per row in channel result CSVs.",
            "artifacts": f"Scores from channel CSVs; model {eff_model_name} ({eff_model_version}).",
        },
        "methodology": {
            "model": f"Hybrid detector ({eff_model_name}); scores are reconstruction/prediction errors (unsupervised).",
            "threshold_source": (
                f"Flag threshold ≈ {thr_used if thr_used is not None else '—'}; "
                f"operating point: {ot.get('name', '—')}. {ot.get('note', '')}"
            ),
            "decision_rule": (
                f"Anomaly if score &gt; threshold. Severity from r = score÷thr: "
                f"Low/Medium/High bands at {SEVERITY_RATIO_MEDIUM_MIN:g} and {SEVERITY_RATIO_HIGH_MIN:g}."
            ),
            "weights_reference": (thresholds_snapshot.get("weights") or {}) if isinstance(thresholds_snapshot, dict) else {},
        },
        "system_overview": {
            "description": "Telemetry anomaly detection via hybrid score vs threshold.",
            "model_behavior": "Deviation score from recon/prediction/gradient terms.",
            "thresholding": "Binary anomaly flag at operating threshold; severity from ratio r.",
        },
        "analysis_meta": meta,
        "detection_results": {
            **summary,
            "model_version": eff_model_version,
            "analysis_timestamp": run.get("created_at"),
            "findings_rows_in_report": findings_in_table,
            "findings_truncated": findings_truncated,
            "findings_cap": cap,
        },
        "channel_summary": channel_summary,
        "risk_classification": {
            "low": summary["low_count"],
            "medium": summary["medium_count"],
            "high": summary["high_count"],
            "logic": {
                "normal": "score ≤ operating threshold (ratio r ≤ 1).",
                "low": f"flagged anomalous: 1 < r < {SEVERITY_RATIO_MEDIUM_MIN:g}",
                "medium": f"{SEVERITY_RATIO_MEDIUM_MIN:g} ≤ r < {SEVERITY_RATIO_HIGH_MIN:g}",
                "high": f"r ≥ {SEVERITY_RATIO_HIGH_MIN:g}",
            },
        },
        "severity_definitions": {
            "basis": "r = score ÷ operating_threshold.",
            "bands": [
                {"name": "Normal", "condition": "r ≤ 1", "meaning": "Below threshold."},
                {"name": "Low", "condition": f"1 < r < {SEVERITY_RATIO_MEDIUM_MIN:g}", "meaning": "Mild."},
                {"name": "Medium", "condition": f"{SEVERITY_RATIO_MEDIUM_MIN:g} ≤ r < {SEVERITY_RATIO_HIGH_MIN:g}", "meaning": "Moderate."},
                {"name": "High", "condition": f"r ≥ {SEVERITY_RATIO_HIGH_MIN:g}", "meaning": "Strong."},
            ],
            "per_finding_fields": "Rows: score, threshold, r, severity from CSV/bands.",
        },
        "mission_risk_policy": {
            "overall_posture": summary["severity_summary"],
            "criteria": [
                "Critical: High window or rate ≥ 10%.",
                "Suspicious: Medium window or rate ≥ 3%.",
                "Normal: else.",
            ],
            "disclaimer": "Advisory only — correlate with mission context.",
        },
        "thresholds_snapshot": thresholds_snapshot,
        "operating_snapshot": operating_snapshot,
        "detailed_findings": [f.__dict__ for f in findings],
        "security_interpretation": interpretation,
        "recommendations": recs,
        "technical_appendix": tech,
        "missing_fields": [
            k
            for k in ("filename", "file_sha256", "created_at")
            if not run.get(k)
        ],
    }
    return report

def write_excel(report: Dict[str, Any], out_path: Path) -> None:
    """
    Power-BI-oriented workbook: no file paths or internal technical fields.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    NA = "Not Available"

    def cell_val(x: Any) -> Any:
        if x is None or (isinstance(x, float) and (x != x)):  # NaN
            return NA
        if isinstance(x, str) and not x.strip():
            return NA
        return x

    def _num(x: Any) -> Optional[float]:
        if x is None:
            return None
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    FILL_RED = PatternFill("solid", fgColor="FECACA")
    FILL_ORANGE = PatternFill("solid", fgColor="FED7AA")
    FILL_YELLOW = PatternFill("solid", fgColor="FEF08A")
    FILL_GREEN = PatternFill("solid", fgColor="BBF7D0")
    FILL_HEADER = PatternFill("solid", fgColor="1E3A8A")
    FONT_HEADER = Font(color="FFFFFF", bold=True)
    FONT_BODY = Font(color="000000")

    def fill_for_posture(s: str) -> PatternFill:
        u = str(s).strip().upper()
        if u == "CRITICAL":
            return FILL_RED
        if u == "SUSPICIOUS":
            return FILL_ORANGE
        return FILL_GREEN

    def fill_for_severity_label(sev: str) -> PatternFill:
        u = str(sev).strip().upper()
        if u in ("HIGH", "CRITICAL"):
            return FILL_RED
        if u in ("MEDIUM", "SUSPICIOUS"):
            return FILL_ORANGE
        if u == "LOW":
            return FILL_YELLOW
        return FILL_GREEN

    def risk_color_label(sev: str) -> str:
        u = str(sev).strip().upper()
        if u in ("HIGH", "CRITICAL"):
            return "Red"
        if u in ("MEDIUM", "SUSPICIOUS"):
            return "Orange"
        if u == "LOW":
            return "Yellow"
        return "Green"

    def decision_from_score(score: Any, thr: Any) -> str:
        s = _num(score)
        t = _num(thr)
        if s is None or t is None:
            return NA
        return "Anomalous" if s > t else "Normal"

    def style_header_row(ws, row: int, ncol: int) -> None:
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def autofit_columns(ws, min_w: float = 10.0, max_w: float = 52.0) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = min_w
            for cell in col:
                v = cell.value
                if v is None:
                    continue
                max_len = max(max_len, min(max_w, len(str(v)) + 2))
            ws.column_dimensions[letter].width = max_len

    def add_excel_table(ws, top_left: str, bottom_right: str, name: str) -> None:
        t = Table(displayName=name, ref=f"{top_left}:{bottom_right}")
        t.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(t)

    cover = report.get("cover") or {}
    es = report.get("executive_summary") or {}
    dr = report.get("detection_results") or {}
    rc = report.get("risk_classification") or {}
    tech = report.get("technical_appendix") or {}
    findings_raw = list(report.get("detailed_findings") or [])
    recs = list(report.get("recommendations") or [])

    report_id = cell_val(tech.get("generated_report_id"))
    run_id = cell_val(cover.get("run_id"))
    generated = cell_val(cover.get("generated_at"))
    uploaded = cell_val(cover.get("telemetry_filename"))
    analyst = cell_val(cover.get("analyst"))
    overall = cell_val(es.get("overall_state") or dr.get("severity_summary"))
    risk_level = overall
    hi_c = dr.get("high_count")
    if hi_c is None:
        hi_c = rc.get("high")
    med_c = dr.get("medium_count")
    if med_c is None:
        med_c = rc.get("medium")
    low_c = dr.get("low_count")
    if low_c is None:
        low_c = rc.get("low")

    tw = dr.get("total_windows")
    nw = dr.get("normal_count")
    aw = dr.get("anomaly_count")
    ar_raw = dr.get("anomaly_rate")
    score_max = dr.get("score_max")
    score_mean = dr.get("score_mean")
    thr_used = dr.get("threshold_used")
    model_ver = dr.get("model_version")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ----- Sheet 1: Summary (Metric | Value) -----
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ("Report ID", report_id),
        ("Run ID", run_id),
        ("Generated Time", generated),
        ("Uploaded File", uploaded),
        ("Analyst", analyst),
        ("Overall Status", overall),
        ("Total Windows", cell_val(tw)),
        ("Normal Windows", cell_val(nw)),
        ("Anomalous Windows", cell_val(aw)),
        ("Anomaly Rate", cell_val(ar_raw) if ar_raw is None else ar_raw),
        ("Highest Score", cell_val(score_max) if score_max is None else score_max),
        ("Mean Score", cell_val(score_mean) if score_mean is None else score_mean),
        ("Threshold Used", cell_val(thr_used) if thr_used is None else thr_used),
        ("Model Version", cell_val(model_ver)),
        ("Risk Level", risk_level),
        ("High Count", cell_val(hi_c) if hi_c is None else hi_c),
        ("Medium Count", cell_val(med_c) if med_c is None else med_c),
        ("Low Count", cell_val(low_c) if low_c is None else low_c),
    ]
    ws.append(["Metric", "Value"])
    for m, v in summary_rows:
        ws.append([m, v])
    style_header_row(ws, 1, 2)
    ws.freeze_panes = "A2"
    posture_fill = fill_for_posture(str(overall))

    for r_idx, (metric, _) in enumerate(summary_rows, start=2):
        val_cell = ws.cell(row=r_idx, column=2)
        val_cell.font = FONT_BODY
        m = str(metric)
        if m in ("Anomaly Rate",) and ar_raw is not None and isinstance(val_cell.value, (int, float)):
            val_cell.number_format = "0.00%"
        elif m in ("Highest Score", "Mean Score") and isinstance(val_cell.value, (int, float)):
            val_cell.number_format = "0.0000"
        elif m == "Threshold Used" and isinstance(val_cell.value, (int, float)):
            val_cell.number_format = "0.00000"
        elif m in ("Total Windows", "Normal Windows", "Anomalous Windows", "High Count", "Medium Count", "Low Count"):
            if isinstance(val_cell.value, (int, float)):
                val_cell.number_format = "0"

        if m in ("Overall Status", "Risk Level"):
            val_cell.fill = posture_fill
        elif m == "High Count":
            val_cell.fill = FILL_RED
        elif m == "Medium Count":
            val_cell.fill = FILL_ORANGE
        elif m == "Low Count":
            val_cell.fill = FILL_YELLOW
        elif m in ("Anomaly Rate", "Highest Score", "Mean Score"):
            val_cell.fill = posture_fill
        elif m == "Threshold Used":
            val_cell.fill = posture_fill

    autofit_columns(ws)
    nsum = len(summary_rows) + 1
    add_excel_table(ws, "A1", f"B{nsum}", "tblSummary")

    # ----- Sheet 2: Findings -----
    ws_f = wb.create_sheet("Findings")
    hdr_f = [
        "Window",
        "Score",
        "Threshold",
        "Ratio",
        "Severity",
        "Channel",
        "Decision",
        "Risk Color Label",
    ]
    ws_f.append(hdr_f)
    for f in findings_raw:
        score = f.get("score")
        thr = f.get("threshold")
        sev = f.get("severity")
        dec = decision_from_score(score, thr)
        rlab = risk_color_label(str(sev or ""))
        win_ix = f.get("window_index")
        ws_f.append(
            [
                cell_val(win_ix),
                cell_val(score) if score is None else score,
                cell_val(thr) if thr is None else thr,
                cell_val(f.get("ratio")) if f.get("ratio") is None else f.get("ratio"),
                cell_val(sev),
                cell_val(f.get("channel")),
                dec,
                rlab,
            ]
        )
    style_header_row(ws_f, 1, len(hdr_f))
    ws_f.freeze_panes = "A2"
    for r in range(2, ws_f.max_row + 1):
        sev_s = str(ws_f.cell(row=r, column=5).value or "")
        row_fill = fill_for_severity_label(sev_s)
        dec_s = str(ws_f.cell(row=r, column=7).value or "")
        if dec_s.upper() == "NORMAL":
            row_fill = FILL_GREEN
        for c in range(1, 9):
            cell = ws_f.cell(row=r, column=c)
            cell.fill = row_fill
            cell.font = FONT_BODY
            if c == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"
            elif c == 3 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00000"
            elif c == 4 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"
    autofit_columns(ws_f)
    if ws_f.max_row > 1:
        add_excel_table(ws_f, "A1", f"{get_column_letter(len(hdr_f))}{ws_f.max_row}", "tblFindings")

    # ----- Sheet 3: Recommendations -----
    ws_r = wb.create_sheet("Recommendations")
    ws_r.append(["Risk Level", "Recommendation"])
    style_header_row(ws_r, 1, 2)
    for text in recs:
        ws_r.append([risk_level, cell_val(text)])
    for r in range(2, ws_r.max_row + 1):
        ws_r.cell(row=r, column=1).fill = posture_fill
        ws_r.cell(row=r, column=1).font = FONT_BODY
        ws_r.cell(row=r, column=2).font = FONT_BODY
        ws_r.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws_r.freeze_panes = "A2"
    autofit_columns(ws_r)
    if ws_r.max_row > 1:
        add_excel_table(ws_r, "A1", f"B{ws_r.max_row}", "tblRecommendations")

    # ----- Sheet 4: PowerBI_Data -----
    ws_p = wb.create_sheet("PowerBI_Data")
    p_hdr = [
        "Report ID",
        "Run ID",
        "Uploaded File",
        "Overall Status",
        "Window",
        "Score",
        "Threshold",
        "Ratio",
        "Severity",
        "Channel",
        "Decision",
        "Anomaly Rate",
        "Risk Level",
        "High Count",
        "Medium Count",
        "Low Count",
    ]
    ws_p.append(p_hdr)
    style_header_row(ws_p, 1, len(p_hdr))

    ar_disp = ar_raw
    hi_d = hi_c if hi_c is not None else NA
    med_d = med_c if med_c is not None else NA
    low_d = low_c if low_c is not None else NA

    if not findings_raw:
        ws_p.append(
            [
                report_id,
                run_id,
                uploaded,
                overall,
                NA,
                NA,
                NA,
                NA,
                NA,
                NA,
                NA,
                ar_disp if ar_disp is not None else NA,
                risk_level,
                hi_d,
                med_d,
                low_d,
            ]
        )
    else:
        for f in findings_raw:
            score = f.get("score")
            thr = f.get("threshold")
            dec = decision_from_score(score, thr)
            ws_p.append(
                [
                    report_id,
                    run_id,
                    uploaded,
                    overall,
                    cell_val(f.get("window_index")),
                    cell_val(score) if score is None else score,
                    cell_val(thr) if thr is None else thr,
                    cell_val(f.get("ratio")) if f.get("ratio") is None else f.get("ratio"),
                    cell_val(f.get("severity")),
                    cell_val(f.get("channel")),
                    dec,
                    ar_disp if ar_disp is not None else NA,
                    risk_level,
                    hi_d,
                    med_d,
                    low_d,
                ]
            )

    ws_p.freeze_panes = "A2"
    posture_fill_p = posture_fill
    for r in range(2, ws_p.max_row + 1):
        sev_s = str(ws_p.cell(row=r, column=9).value or "")
        row_fill = fill_for_severity_label(sev_s)
        if str(ws_p.cell(row=r, column=11).value or "").upper() == "NORMAL":
            row_fill = FILL_GREEN
        for c in range(1, len(p_hdr) + 1):
            cell = ws_p.cell(row=r, column=c)
            cell.fill = row_fill
            cell.font = FONT_BODY
            if c == 6 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"
            elif c == 7 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00000"
            elif c == 8 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"
            elif c == 12 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
        ws_p.cell(row=r, column=4).fill = posture_fill_p
        ws_p.cell(row=r, column=13).fill = posture_fill_p
        ws_p.cell(row=r, column=14).fill = FILL_RED
        ws_p.cell(row=r, column=15).fill = FILL_ORANGE
        ws_p.cell(row=r, column=16).fill = FILL_YELLOW

    autofit_columns(ws_p)
    if ws_p.max_row > 1:
        add_excel_table(
            ws_p,
            "A1",
            f"{get_column_letter(len(p_hdr))}{ws_p.max_row}",
            "tblPowerBIData",
        )

    wb.save(out_path)

def write_pdf(report: Dict[str, Any], out_path: Path) -> None:
    """
    Ultra-compact PDF (target 2–3 pages): cover, executive KPIs, risk line, top 10 anomalies,
    up to 3 recommendation lines, minimal technical note. Does not alter JSON/Excel data.
    """
    import re

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.pdfgen.canvas import Canvas

    PDF_TOP_ANOMALIES = 10
    PDF_MAX_RECS = 3

    out_path.parent.mkdir(parents=True, exist_ok=True)

    report_id = report.get("technical_appendix", {}).get("generated_report_id", "—")
    W = A4[0]

    C_NAVY = colors.HexColor("#0B2A4A")
    C_BLUE = colors.HexColor("#1E3A8A")
    C_SLATE = colors.HexColor("#334155")
    C_WHITE = colors.white
    C_GREY_RULE = colors.HexColor("#CBD5E1")
    C_GREY_ROW = colors.HexColor("#F1F5F9")
    C_RED = colors.HexColor("#DC2626")
    C_ORANGE = colors.HexColor("#EA580C")
    C_GREEN = colors.HexColor("#16A34A")

    styles = getSampleStyleSheet()

    def _ps(name: str, **kw: Any) -> ParagraphStyle:
        base = kw.pop("parent", styles["Normal"])
        return ParagraphStyle(name, parent=base, **kw)

    title_style = _ps(
        "pdf_title",
        fontSize=16,
        leading=19,
        alignment=TA_CENTER,
        textColor=C_NAVY,
        fontName="Helvetica-Bold",
        spaceAfter=8,
    )
    h_bar = _ps("pdf_hbar", fontSize=10, leading=12, fontName="Helvetica-Bold", textColor=C_WHITE)
    body = _ps("pdf_body", fontSize=9, leading=11, textColor=C_NAVY)
    small = _ps("pdf_small", fontSize=8, leading=10, textColor=C_SLATE)
    tbl_hdr = _ps("pdf_th", fontSize=7.5, fontName="Helvetica-Bold", alignment=TA_CENTER, textColor=C_WHITE)
    tbl_cell = _ps("pdf_tc", fontSize=7.5, leading=9, textColor=C_NAVY)
    mono = _ps("pdf_mono", fontSize=7.5, fontName="Courier", textColor=C_NAVY)

    def _strip_one_line_reason(raw: Any, max_len: int = 130) -> str:
        t = re.sub(r"<[^>]+>", "", str(raw or ""))
        t = " ".join(t.split())
        if len(t) > max_len:
            return t[: max_len - 1] + "…"
        return t or "—"

    def _section_bar(label: str) -> Table:
        return Table(
            [[Paragraph(label, h_bar)]],
            colWidths=[174 * mm],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), C_BLUE),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]),
        )

    def _kv_compact(rows: List[Tuple[str, Any]], fs: float = 8.5) -> Table:
        data = []
        for i, (k, v) in enumerate(rows):
            data.append(
                [
                    Paragraph(f"<b>{k}</b>", _ps(f"kk_{i}", fontSize=fs, textColor=C_NAVY, fontName="Helvetica-Bold")),
                    Paragraph(str(v), _ps(f"vv_{i}", fontSize=fs, textColor=C_SLATE)),
                ]
            )
        t = Table(data, colWidths=[52 * mm, 122 * mm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EFF6FF")),
                    ("GRID", (0, 0), (-1, -1), 0.25, C_GREY_RULE),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        return t

    cover = report.get("cover", {})
    es = report.get("executive_summary", {})
    dr = report.get("detection_results", {})
    rc = report.get("risk_classification", {})
    tech = report.get("technical_appendix", {}) or {}
    meth = report.get("methodology", {})

    posture_raw = str(es.get("overall_state") or dr.get("severity_summary") or "Normal").strip()
    pu = posture_raw.upper()
    status_color = C_GREEN
    status_bg = colors.HexColor("#F0FDF4")
    if pu == "SUSPICIOUS":
        status_color = C_ORANGE
        status_bg = colors.HexColor("#FFFBEB")
    elif pu == "CRITICAL":
        status_color = C_RED
        status_bg = colors.HexColor("#FEF2F2")

    def _footer(canvas: Canvas, doc: Any) -> None:
        canvas.saveState()
        canvas.setStrokeColor(C_GREY_RULE)
        canvas.setLineWidth(0.35)
        canvas.line(18 * mm, 16 * mm, W - 18 * mm, 16 * mm)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(C_SLATE)
        canvas.drawString(18 * mm, 10 * mm, "Confidential")
        canvas.drawRightString(W - 18 * mm, 10 * mm, f"Page {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=14 * mm,
        bottomMargin=18 * mm,
        title="CyberSatDetect Security Report",
        author="CyberSatDetect",
    )

    story: list = []

    # --- A) Cover / header ---
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("CyberSatDetect Security Report", title_style))
    story.append(
        _kv_compact(
            [
                ("Report ID", report_id),
                ("Run ID", cover.get("run_id") or "—"),
                ("Generated Time (UTC)", cover.get("generated_at") or "—"),
                ("Uploaded File", cover.get("telemetry_filename") or "—"),
                ("Analyst", cover.get("analyst") or "—"),
            ],
            fs=9,
        )
    )
    story.append(Spacer(1, 4 * mm))
    st_tbl = Table(
        [[Paragraph(f"<b>Overall Status: {posture_raw}</b>", _ps("st", fontSize=11, textColor=status_color, fontName="Helvetica-Bold", alignment=TA_CENTER))]],
        colWidths=[174 * mm],
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), status_bg),
                ("BOX", (0, 0), (-1, -1), 1.2, status_color),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        ),
    )
    story.append(st_tbl)
    story.append(Spacer(1, 6 * mm))

    # --- B) Executive summary ---
    story.append(_section_bar("Executive Summary"))
    story.append(Spacer(1, 3 * mm))
    ar_pct = float(dr.get("anomaly_rate") or 0) * 100.0
    exec_rows = [
        ("Total Windows", dr.get("total_windows", "—")),
        ("Normal Windows", dr.get("normal_count", "—")),
        ("Anomalous Windows", dr.get("anomaly_count", "—")),
        ("Anomaly Rate", f"{ar_pct:.2f}%"),
        ("Highest Score", dr.get("score_max") if dr.get("score_max") is not None else "—"),
        ("Mean Score", dr.get("score_mean") if dr.get("score_mean") is not None else "—"),
        ("Threshold Used", dr.get("threshold_used") if dr.get("threshold_used") is not None else "—"),
        ("Model Version", dr.get("model_version") or "—"),
    ]
    story.append(_kv_compact(exec_rows))
    story.append(Spacer(1, 6 * mm))

    # --- C) Risk summary ---
    story.append(_section_bar("Risk Summary"))
    story.append(Spacer(1, 3 * mm))
    risk_rows = [
        ("Risk Level", posture_raw),
        ("High Count", rc.get("high", "—")),
        ("Medium Count", rc.get("medium", "—")),
        ("Low Count", rc.get("low", "—")),
        (
            "Classification rationale (one line)",
            _strip_one_line_reason(es.get("mission_file_assessment"), max_len=300),
        ),
    ]
    story.append(_kv_compact(risk_rows))
    story.append(Spacer(1, 6 * mm))

    # --- D) Top anomalies (10 max) ---
    story.append(_section_bar("Top Anomalies"))
    story.append(Spacer(1, 2 * mm))
    findings_all = report.get("detailed_findings") or []
    findings = findings_all[:PDF_TOP_ANOMALIES]
    hdr = [
        Paragraph("Window", tbl_hdr),
        Paragraph("Score", tbl_hdr),
        Paragraph("Threshold", tbl_hdr),
        Paragraph("Ratio", tbl_hdr),
        Paragraph("Severity", tbl_hdr),
        Paragraph("Channel", tbl_hdr),
    ]
    rows_o = [hdr]
    for ix, f in enumerate(findings):
        sev = str(f.get("severity") or "")
        rows_o.append(
            [
                Paragraph(str(f.get("window_index", "")), tbl_cell),
                Paragraph(f"{float(f.get('score') or 0):.4f}", mono),
                Paragraph(f"{float(f.get('threshold') or 0):.5f}", mono),
                Paragraph(f"{float(f.get('ratio') or 0):.3f}", mono),
                Paragraph(f"<b>{sev.upper()}</b>", tbl_cell),
                Paragraph(str(f.get("channel", ""))[:36], tbl_cell),
            ]
        )
    if len(rows_o) == 1:
        story.append(Paragraph("No anomalous windows in this run.", body))
    else:
        tw = Table(
            rows_o,
            colWidths=[16 * mm, 26 * mm, 28 * mm, 22 * mm, 22 * mm, 60 * mm],
            repeatRows=1,
        )
        tw.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), C_BLUE),
                    ("GRID", (0, 0), (-1, -1), 0.25, C_GREY_RULE),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_GREY_ROW]),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(tw)
    story.append(Spacer(1, 6 * mm))

    # --- E) Recommendations (max 3, one line each) ---
    story.append(_section_bar("Recommendations"))
    story.append(Spacer(1, 2 * mm))
    recs = (report.get("recommendations") or [])[:PDF_MAX_RECS]
    for ri, rec in enumerate(recs):
        line = " ".join(str(rec).split())
        if len(line) > 220:
            line = line[:217] + "…"
        story.append(Paragraph(f"{ri + 1}. {line}", body))
    if not recs:
        story.append(Paragraph("—", small))
    story.append(Spacer(1, 6 * mm))

    # --- F) Technical (brief) ---
    story.append(_section_bar("Technical Info"))
    story.append(Spacer(1, 3 * mm))
    w_str = tech.get("weights")
    if isinstance(w_str, dict):
        w_str = str(w_str)
    tech_rows = [
        ("Model Name", tech.get("model_name") or "—"),
        ("Threshold Strategy", (tech.get("threshold_strategy") or meth.get("threshold_source") or "—")),
        ("Weights", w_str if w_str is not None else "—"),
        ("Report Generated By", "CyberSatDetect"),
    ]
    story.append(_kv_compact(tech_rows))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)

